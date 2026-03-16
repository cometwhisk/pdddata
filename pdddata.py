import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import re
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

class TimeEntryGroup:
    """终极时间输入组件：全空初始态、Tab 补零、全选、自动跳格"""
    def __init__(self, parent):
        self.frame = tk.Frame(parent, bg="#ffffff")
        
        # 创建三个输入框，默认全空
        self.h_ent = self._create_box(self.frame)
        tk.Label(self.frame, text=":", bg="#ffffff").pack(side="left")
        self.m_ent = self._create_box(self.frame)
        tk.Label(self.frame, text=":", bg="#ffffff").pack(side="left")
        self.s_ent = self._create_box(self.frame)
        
        # 绑定事件
        for ent, next_ent in [(self.h_ent, self.m_ent), (self.m_ent, self.s_ent), (self.s_ent, None)]:
            # 获得焦点自动全选，方便直接覆盖输入
            ent.bind("<FocusIn>", lambda e, x=ent: x.selection_range(0, tk.END))
            # 失去焦点自动补零（例如输 1 变 01）
            ent.bind("<FocusOut>", lambda e, x=ent: self._format_on_leave(x))
            # 输入监控与自动跳格
            ent.bind("<KeyRelease>", lambda e, cur=ent, nxt=next_ent: self._on_key(e, cur, nxt))

    def _create_box(self, parent):
        # 移除了边框，采用扁平化设计
        ent = tk.Entry(parent, width=3, justify="center", font=('Segoe UI', 10), 
                      relief="flat", highlightthickness=1, highlightbackground="#cccccc")
        ent.pack(side="left", padx=2, pady=2)
        return ent

    def _format_on_leave(self, current):
        val = current.get().strip()
        if val and len(val) == 1:
            current.delete(0, tk.END)
            current.insert(0, val.zfill(2))

    def _on_key(self, event, current, next_box):
        if event.keysym in ("Tab", "Shift_L", "Shift_R", "BackSpace", "Left", "Right"): 
            return
        
        val = current.get()
        # 输入冒号时自动补全并跳转
        if event.char in (":", "："):
            current.delete(0, tk.END)
            current.insert(0, val.replace(":", "").replace("：", "").zfill(2))
            if next_box: next_box.focus_set()
            return

        clean_val = re.sub(r'\D', '', val)
        if val != clean_val:
            current.delete(0, tk.END)
            current.insert(0, clean_val)

        # 输入满2位自动跳下一格
        if len(clean_val) >= 2:
            current.delete(0, tk.END)
            current.insert(0, clean_val[:2])
            if next_box: next_box.focus_set()

    def get_time(self):
        # 如果用户没填，自动补 00，并检查数值范围
        h = self.h_ent.get().strip().zfill(2)
        m = self.m_ent.get().strip().zfill(2)
        s = self.s_ent.get().strip().zfill(2)
        
        if not (0 <= int(h or 0) <= 23 and 0 <= int(m or 0) <= 59 and 0 <= int(s or 0) <= 59):
            raise ValueError("时间数值范围不合法")
        return f"{h}:{m}:{s}"
    
    def clear(self):
        """清空小时、分钟、秒的输入框"""
        for ent in [self.h_ent, self.m_ent, self.s_ent]:
            ent.delete(0, tk.END)  # 从第0位删到最后一位

class SalesApp:
    def __init__(self, root):
        self.root = root
        self.root.title("拼多多单量分析助手 V2.1")
        self.root.configure(bg="#f5f5f5")
        
        # 窗口居中
        w, h = 850, 800
        x = (self.root.winfo_screenwidth() - w) // 2
        y = (self.root.winfo_screenheight() - h) // 2
        self.root.geometry(f"{w}x{h}+{x}+{y}")

        self.blacklist_list = []
        self.file_path = None
        
        # 严格匹配 aa.csv 的原始表头顺序
        self.STANDARD_HEADER = [
            '商品', '订单号', '订单状态', '商品总价(元)', '邮费(元)', '店铺优惠折扣(元)', 
            '平台优惠折扣(元)', '多多支付立减金额(元)', '用户实付金额(元)', '商家实收金额(元)', 
            '商品数量(件)', '发货时间', '确认收货时间', '商品id', '商品规格', '样式ID', 
            '商家编码-规格维度', '商家编码-商品维度', '商家备注', '售后状态', '快递单号', 
            '快递公司', '订单成交时间', '是否分期', '分期期数', '手续费承担方', '分期方式'
        ]

        self.style = ttk.Style()
        self.style.theme_use('clam')
        self.style.configure("Main.TButton", font=("Microsoft YaHei", 10, "bold"), background="#0078d4", foreground="white")
        self.setup_ui()

    def setup_ui(self):
        # 蓝色页眉
        top = tk.Frame(self.root, bg="#0078d4", height=60); top.pack(fill="x")
        tk.Label(top, text="拼多多单量分析助手", bg="#0078d4", fg="white", font=("Microsoft YaHei", 16, "bold")).pack(pady=15)

        container = tk.Frame(self.root, bg="#f5f5f5"); container.pack(fill="both", expand=True, padx=30, pady=10)

        # 1. 载入
        f_box = tk.LabelFrame(container, text=" 第一步：载入 CSV 订单表 ", bg="#f5f5f5", font=("Microsoft YaHei", 10, "bold"), padx=10, pady=10)
        f_box.pack(fill="x", pady=10)
        ttk.Button(f_box, text="📁 选择文件", command=self.load_file, width=20).pack(side="left")
        self.file_label = tk.Label(f_box, text="等待载入...", bg="#f5f5f5", fg="#666666"); self.file_label.pack(side="left", padx=10)

        # 2. 排除（两个时间框现在都是默认空的）
        ex_box = tk.LabelFrame(container, text=" 第二步：排除干扰区间 (选填) ", bg="#f5f5f5", font=("Microsoft YaHei", 10, "bold"), padx=10, pady=10)
        ex_box.pack(fill="x", pady=10)
        row1 = tk.Frame(ex_box, bg="#f5f5f5"); row1.pack(fill="x")
        
        tk.Label(row1, text="从:", bg="#f5f5f5").grid(row=0, column=0)
        self.date_s = ttk.Combobox(row1, state="readonly", width=12); self.date_s.grid(row=0, column=1, padx=5)
        self.time_s = TimeEntryGroup(row1)
        self.time_s.frame.grid(row=0, column=2)

        tk.Label(row1, text=" 至 ", bg="#f5f5f5").grid(row=0, column=3, padx=5)
        self.date_e = ttk.Combobox(row1, state="readonly", width=12); self.date_e.grid(row=0, column=4, padx=5)
        self.time_e = TimeEntryGroup(row1) # 修正：这里也默认空
        self.time_e.frame.grid(row=0, column=5)

        ttk.Button(row1, text="添加排除", command=self.add_blacklist).grid(row=0, column=6, padx=15)
        self.blacklist_box = tk.Listbox(ex_box, height=4, font=('Consolas', 9), relief="flat", highlightthickness=1, highlightbackground="#dddddd")
        self.blacklist_box.pack(fill="x", pady=10)
        ttk.Button(ex_box, text="删除选中项", command=self.remove_blacklist).pack(side="right")

        # 3. 对比
        d_box = tk.LabelFrame(container, text=" 第三步：对比分析日期 (按 ID 汇总) ", bg="#f5f5f5", font=("Microsoft YaHei", 10, "bold"), padx=10, pady=10)
        d_box.pack(fill="x", pady=10)
        self.date_a = ttk.Combobox(d_box, state="readonly", width=18); self.date_a.pack(side="left", padx=20)
        tk.Label(d_box, text="VS", bg="#f5f5f5", font=("Arial", 10, "bold")).pack(side="left")
        self.date_b = ttk.Combobox(d_box, state="readonly", width=18); self.date_b.pack(side="left", padx=20)

        # 4. 执行
        self.run_btn = ttk.Button(self.root, text="🚀 开始分析并导出结果", style="Main.TButton", command=self.process_data)
        self.run_btn.pack(pady=20, ipady=10, ipadx=20)

    def reset_ui(self):
        """核心重置逻辑：恢复到初始状态"""
        # 1. 清空数据变量
        self.file_path = None
        self.blacklist_list = []
        
        # 2. 恢复文字提示
        self.file_label.config(text="等待载入...", fg="#666666")
        
        # 3. 清空排除区间的列表
        self.blacklist_box.delete(0, tk.END)
        
        # 4. 调用刚才写好的时间清空
        self.time_s.clear()
        self.time_e.clear()
        
        # 5. 清空并复原所有日期下拉框
        for c in [self.date_s, self.date_e, self.date_a, self.date_b]:
            c['values'] = [] # 清空可选日期
            c.set('')        # 变成空白

    def load_file(self):
        p = filedialog.askopenfilename(filetypes=[("CSV Files", "*.csv")])
        if not p: return
        try:
            # 严格校验顺序和内容
            actual_header = pd.read_csv(p, nrows=0).columns.tolist()
            if actual_header != self.STANDARD_HEADER:
                messagebox.showerror("格式错误", "表格列名或顺序与标准不符，请检查导出方式！")
                return

            df = pd.read_csv(p)
            df['订单成交时间'] = pd.to_datetime(df['订单成交时间'].str.strip(), errors='coerce')
            df = df.dropna(subset=['订单成交时间'])
            dates = sorted(df['订单成交时间'].dt.date.unique().astype(str))
            for c in [self.date_s, self.date_e, self.date_a, self.date_b]:
                c['values'] = dates
                if dates: c.set(dates[0])
            self.file_path = p
            self.file_label.config(text=f"已载入: {p.split('/')[-1]}", fg="#0078d4")
            messagebox.showinfo("导入成功", f"文件校验通过！\n本次共成功导入 {len(df)} 条订单数据。")
        except Exception as e:
            messagebox.showerror("错误", str(e))

    def add_blacklist(self):
        if not self.file_path: return
        try:
            fs = f"{self.date_s.get()} {self.time_s.get_time()}"
            fe = f"{self.date_e.get()} {self.time_e.get_time()}"
            if pd.to_datetime(fe) <= pd.to_datetime(fs):
                messagebox.showwarning("提醒", "结束时间需晚于开始时间")
                return
            self.blacklist_list.append((fs, fe))
            self.blacklist_box.insert(tk.END, f" 排除: {fs} 至 {fe}")
        except Exception as e: messagebox.showerror("格式错误", str(e))

    def remove_blacklist(self):
        sel = self.blacklist_box.curselection()
        if sel: idx = sel[0]; self.blacklist_box.delete(idx); self.blacklist_list.pop(idx)

    def process_data(self):
        if not self.file_path: return
        da, db = self.date_a.get(), self.date_b.get()
        if not da or not db: return
        try:
            df = pd.read_csv(self.file_path)
            df = df[~df['订单状态'].isin(['待付款', '已取消'])]
            df['订单成交时间'] = pd.to_datetime(df['订单成交时间'].str.strip(), errors='coerce')
            df = df.dropna(subset=['订单成交时间'])
            
            # 过滤黑名单
            for s, e in self.blacklist_list:
                df = df[~((df['订单成交时间'] >= pd.to_datetime(s)) & (df['订单成交时间'] <= pd.to_datetime(e)))]
            
            df['成交日期'] = df['订单成交时间'].dt.date.astype(str)
            # 先去除可能存在的空格，再转为数字类型（int64 适合存储长 ID）
            df['商品id'] = pd.to_numeric(df['商品id'], errors='coerce').fillna(0).astype('int64')
            
            # 汇总对比
            daily = df.groupby(['成交日期', '商品id']).size().reset_index(name='销量')
            ca, cb = f'{da}', f'{db}'
            sa = daily[daily['成交日期']==da][['商品id','销量']].rename(columns={'销量':ca})
            sb = daily[daily['成交日期']==db][['商品id','销量']].rename(columns={'销量':cb})
            
            res = pd.merge(sa, sb, on='商品id', how='outer').fillna(0)
            res['单差'] = (res[cb] - res[ca]).astype(int)
            res = res.sort_values(by='单差', ascending=True)
            
            sp = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile=f"单量分析_{da}_{db}.xlsx")
            if sp: 
                with pd.ExcelWriter(sp, engine='openpyxl') as writer:
                    res.to_excel(writer, index=False, sheet_name='分析结果')
                    worksheet = writer.book['分析结果']
                    
                    # --- 1. 定义样式（像选衣服一样定义好样式） ---
                    # 表头样式：深蓝色背景，白色粗体字，居中
                    header_fill = PatternFill(start_color="0078D4", end_color="0078D4", fill_type="solid")
                    header_font = Font(name='微软雅黑',color="FFFFFF", bold=True, size=11)
                    body_font = Font(name='微软雅黑', size=10)
                    center_alignment = Alignment(horizontal="center", vertical="center")
                    
                    # 边框样式：细实线
                    thin_side = Side(style="thin", color="000000")
                    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

                    # --- 2. 应用样式到表头 ---
                    for cell in worksheet[1]: # 第1行是表头
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = center_alignment
                        cell.border = thin_border
                    # 设置第一行（表头）的高度为 30（默认通常是 15 左右）
                    worksheet.row_dimensions[1].height = 30

                    # --- 3. 自动调整列宽 + 数据行居中 + 加边框 ---
                    for i, col in enumerate(res.columns):
                        # 计算这一列最长的内容长度（加上一点余量）
                        max_len = max(res[col].astype(str).map(len).max(), len(col)) + 4
                        worksheet.column_dimensions[chr(65 + i)].width = max_len
                        
                        # 给数据行加边框和居中
                        for row in range(2, len(res) + 2):
                            cell = worksheet.cell(row=row, column=i+1)
                            cell.alignment = center_alignment
                            cell.border = thin_border

                            cell.font = body_font

                            #针对 ID 列设置数字格式
                            if col == '商品id':
                                cell.number_format = '0' # 强制显示为完整数字
                            
                            # --- 4. 业务逻辑高亮（灵魂点睛） ---
                            # 如果是“单差”这一列（最后一列），根据正负标颜色
                            if col == '单差':
                                if cell.value < 0:
                                    cell.font = Font(name='微软雅黑',color="FF0000", bold=True) # 掉量了，红色警告
                                elif cell.value > 0:
                                    cell.font = Font(name='微软雅黑',color="00B050", bold=True) # 涨了，绿色鼓励

                    # 5. 冻结首行
                    worksheet.freeze_panes = 'A2'
                
                messagebox.showinfo("成功", "分析报告已生成！")
                self.reset_ui()
        except Exception as e: messagebox.showerror("故障", str(e))

if __name__ == "__main__":
    root = tk.Tk(); SalesApp(root); root.mainloop()